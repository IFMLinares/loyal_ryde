import calendar
import locale
from babel.dates import format_date
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.db.models import Count
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.template.loader import render_to_string
from django.urls import reverse_lazy, reverse
from django.utils.html import strip_tags
from django.utils.timezone import now
from django.views.generic import (
    CreateView, 
    DeleteView, 
    DetailView, 
    ListView,
    TemplateView, 
    UpdateView, 
    View,
    )
from django.conf import settings
from apps.loyalRyde.forms import *
from apps.loyalRyde.models import *


# Agregar nuevo traslado
class TransferRequestCreateView(LoginRequiredMixin, CreateView):
    model = TransferRequest
    template_name = 'loyal_ryde_system/transfer_rerquest.html'
    form_class = TransferRequestForm
    success_url = reverse_lazy('core:transfer_request_list')

    def send_email_creation_transfer(self, transfer_request):
        # Recuperar los usuarios con los roles 'administrador' y 'despachador'
        recipients = CustomUser.objects.filter(role__in=['administrador', 'despachador']).values_list('email', flat=True)
        
        subject = 'Nueva solicitud de traslado'
        
        # Construir la URL completa de la imagen
        request = self.request
        image_url = request.build_absolute_uri(settings.STATIC_URL + 'assets/media/logos/logo-01.png')
        
        html_message = render_to_string('emails/test.html', {
            'transfer_request': transfer_request,
            'image_url': image_url
        })
        plain_message = strip_tags(html_message)
        from_email = 'loyalride.test@gmail.com'

        send_mail(subject, plain_message, from_email, recipients, html_message=html_message)

    def form_valid(self, form):
        print("Form is valid, proceeding to save TransferRequest...")  # Depuración
        print(f"Form data: {form.cleaned_data}")  # Depuración
        form.instance.service_requested = self.request.user

        # Asigna el valor de company según el rol del usuario
        if self.request.user.role == "administrador":
            form.instance.company = form.cleaned_data.get('company')
        else:
            form.instance.company = self.request.user.company

        print(f"Company assigned: {form.instance.company}")  # Depuración

        # Guardar el discount_code si viene en el formulario
        discount_code_value = self.request.POST.get('discount_code')
        discount_coupon = None
        if discount_code_value:
            discount_coupon = DiscountCoupon.objects.filter(code=discount_code_value).first()
            if discount_coupon:
                form.instance.discount_coupon = discount_coupon

        transfer_request = form.save()

        # Obtén el ID de la tarifa y el objeto Rates
        rate_id = form.cleaned_data['rate'].id
        rate = Rates.objects.get(id=rate_id)

        # Calcula el precio basado en si es ida y vuelta
        if form.cleaned_data['is_round_trip']:
            transfer_request.price = rate.price_round_trip
        else:
            transfer_request.price = rate.price

        # Calcula el precio final basado en los desvíos
        waypoints_numbers = form.cleaned_data.get('waypoints_numbers', 0)
        if waypoints_numbers > 0:
            transfer_request.final_price += (waypoints_numbers * rate.detour_local)

        # Guarda el objeto TransferRequest con los nuevos valores
        transfer_request.save()

        # Enviar correo electrónico
        self.send_email_creation_transfer(transfer_request)

        return super().form_valid(form)

    def post(self, request, *args, **kwargs):
        # Obtén la fecha directamente del POST
        fecha = request.POST.get('date')
        print(fecha)
        discount_code = self.request.POST.get('discount_code')
        discount_code = DiscountCoupon.objects.filter(code=discount_code).first()

        # Convierte la fecha al formato que Django espera
        # Permite tanto 'YYYY-MM-DD' como 'MM/DD/YYYY'
        try:
            # Si ya está en formato 'YYYY-MM-DD', úsalo directamente
            fecha_obj = datetime.strptime(fecha, '%Y-%m-%d')
        except ValueError:
            # Si no, intenta 'MM/DD/YYYY'
            fecha_obj = datetime.strptime(fecha, '%m/%d/%Y')
        fecha = fecha_obj.strftime('%Y-%m-%d')
        # Actualiza la fecha en los datos del POST
        request.POST = request.POST.copy()
        request.POST['date'] = fecha

        # Llama al método post original para guardar el TransferRequest
        response = super().post(request, *args, **kwargs)

        # Obtén el objeto TransferRequest recién creado
        transfer_request = self.object

        try:
            waypoints_numbers = int(request.POST.get('waypoints_numbers', 0))
            for i in range(3, 3 + waypoints_numbers):
                name = request.POST.get(f'waypoint-{i}')
                lat = request.POST.get(f'lat_{i}')
                lng = request.POST.get(f'lng_{i}')
                if lat and lng:
                    desviation = Desviation.objects.create(
                        desviation_direc=name,
                        desviation_number=i - 2,
                        waypoint_number=i,
                        lat=lat,
                        long=lng
                    )
                    transfer_request.deviation.add(desviation)
        except:
            pass
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        departure_points = DeparturePoint.objects.all()
        context['rates_list'] = Rates.objects.all()
        context['departure'] = departure_points
        return context

# Actualizar traslado
class TransferRequestUpdateView(LoginRequiredMixin, UpdateView):
    model = TransferRequest
    template_name = 'loyal_ryde_system/transfer_rerquest_update.html'
    form_class = TransferRequestForm
    success_url = reverse_lazy('core:transfer_request_list')

    def form_valid(self, form):
        # obtener el id del TransferRequest
        transfer_request_id = self.kwargs.get('pk')
        # obtener el objeto TransferRequest
        transfer_request = TransferRequest.objects.get(id=transfer_request_id)
        # obtener el usuario que tiene en el campo service_requested
        user = transfer_request.service_requested
        # guardar el formulario manteniendo el service_requested
        form.instance.service_requested = user
        # form.instance.service_requested = self.request.user
        messages.success(self.request, 'Formulario guardado exitosamente!')
        return super().form_valid(form)

    def post(self, request, *args, **kwargs):
        # Obtén la fecha directamente del POST
        fecha = request.POST.get('date')
        print(request.POST)
        # Actualiza la fecha en los datos del POST
        request.POST = request.POST.copy()
        request.POST['date'] = fecha
        
        # Llama al método post original para guardar el TransferRequest
        response = super().post(request, *args, **kwargs)

        # Obtén el objeto TransferRequest recién creado
        transfer_request = self.object

        # Procesa los desvíos adicionales
        try:
            waypoints_numbers = int(request.POST.get('waypoints_numbers', 0))
        except ValueError:
            waypoints_numbers = 0
            print("Error: waypoints_numbers no es un entero válido")

        for i in range(3, 3 + waypoints_numbers):
            lat = request.POST.get(f'lat_{i}')
            lng = request.POST.get(f'lng_{i}')
            if lat and lng:
                try:
                    desviation = Desviation.objects.create(
                        desviation_number=i - 2,
                        waypoint_number=i,
                        lat=lat,
                        long=lng
                    )
                    transfer_request.deviation.add(desviation)
                except Exception as e:
                    print(f"Error al crear desvío {i - 2}: {e}")
            else:
                print(f"Latitud o longitud faltante para el desvío {i - 2}")
        return response
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        if self.request.method == 'POST':
            person_ids = self.request.POST.getlist('person_to_transfer')
            if person_ids:
                form.fields['person_to_transfer'].queryset = PeopleTransfer.objects.filter(pk__in=person_ids)
        return form
    
    def get_context_data(self,**kwargs):
        context = super().get_context_data(**kwargs)
        departure_points = DeparturePoint.objects.all()
        context['departure'] = departure_points
        context['desviations'] = self.object.deviation.all()
        return context

# agregar traslado (modo invitado)
class GuestTransferCreateView(CreateView):
    model = TransferRequest
    template_name = 'loyal_ryde_system/transfer_rerquest_guest.html'
    form_class = TransferRequestForm

    def send_email_creation_transfer_guest(self, transfer_request):
        # Recuperar los usuarios con los roles 'administrador' y 'despachador'
        recipients = CustomUser.objects.filter(role__in=['administrador', 'despachador']).values_list('email', flat=True)
        
        subject = 'Nueva solicitud de traslado'
        
        # Construir la URL completa de la imagen
        request = self.request
        image_url = request.build_absolute_uri(settings.STATIC_URL + 'assets/media/logos/logo-01.png')
        
        html_message = render_to_string('emails/test.html', {
            'transfer_request': transfer_request,
            'image_url': image_url
        })
        plain_message = strip_tags(html_message)
        from_email = 'loyalride.test@gmail.com'

        send_mail(subject, plain_message, from_email, recipients, html_message=html_message)

    def form_valid(self, form):
        # Guarda el formulario directamente
        form.instance.service_requested = None
        form.instance.company = None
        
        transfer_request = form.save()

        # Obtén el ID de la tarifa y el objeto Rates
        rate_id = form.cleaned_data['rate'].id
        rate = Rates.objects.get(id=rate_id)

        # Calcula el precio basado en si es ida y vuelta
        if form.cleaned_data['is_round_trip']:
            transfer_request.price = rate.price_round_trip
        else:
            transfer_request.price = rate.price

        # Calcula el precio final basado en los desvíos
        waypoints_numbers = form.cleaned_data.get('waypoints_numbers', 0)
        if waypoints_numbers > 0:
            transfer_request.final_price += (waypoints_numbers * rate.detour_local)

        # Guarda el objeto TransferRequest con los nuevos valores
        transfer_request.save()
        
        self.send_email_creation_transfer_guest(transfer_request)

        return HttpResponseRedirect(reverse('core:guest_transfer_success', kwargs={'pk': transfer_request.pk}))

    def post(self, request, *args, **kwargs):
        # Obtén la fecha directamente del POST
        print(request.POST)
        fecha = request.POST.get('date')

        # Convierte la fecha al formato que Django espera
        fecha = datetime.strptime(fecha, '%m/%d/%Y').strftime('%Y-%m-%d')
        # Actualiza la fecha en los datos del POST
        request.POST = request.POST.copy()
        request.POST['date'] = fecha

        # Llama al método post original para guardar el TransferRequest
        response = super().post(request, *args, **kwargs)

        # Obtén el objeto TransferRequest recién creado
        transfer_request = self.object
        # Procesa los desvíos adicionales

        try:
            waypoints_numbers = int(request.POST.get('waypoints_numbers', 0))
            for i in range(3, 3 + waypoints_numbers):
                name = request.POST.get(f'waypoint-{i}')
                lat = request.POST.get(f'lat_{i}')
                lng = request.POST.get(f'lng_{i}')
                if lat and lng:
                    desviation = Desviation.objects.create(
                        desviation_direc=name,
                        desviation_number=i - 2,
                        waypoint_number=i,
                        lat=lat,
                        long=lng
                    )
                    transfer_request.deviation.add(desviation)
        except:
            pass
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        departure_points = DeparturePoint.objects.all()
        context['rates_list'] = Rates.objects.all()
        context['departure'] = departure_points
        return context

# Detalles del traslado
class TransferRequestDetailview(LoginRequiredMixin, DetailView):
    model = TransferRequest
    template_name = 'loyal_ryde_system/transfer_request_detail.html'
    context_object_name = 'detail'

    def get_context_data(self,**kwargs):
        context = super().get_context_data(**kwargs)
        departure_points = DeparturePoint.objects.all()
        context['departure'] = departure_points
        context['desviations'] = self.object.deviation.all()
        return context

# Añadir traslado modo invitado
class GuestTransferSuccessView(TemplateView):
    template_name = 'loyal_ryde_system/guest_transfer_success.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        transfer_request_id = self.kwargs.get('pk')
        transfer_request = get_object_or_404(TransferRequest, pk=transfer_request_id)
        context['transfer_request'] = transfer_request
        return context

# listado de Traslados
class TransferRequestListView(LoginRequiredMixin, ListView):
    model = TransferRequest
    template_name = 'loyal_ryde_system/transfer_request_list.html'
    context_object_name = 'transfer_requests'
    
    def get_queryset(self):
        user = self.request.user
        if user.is_superuser or (user.company and user.company.name == "Loyal Ride"):
            return TransferRequest.objects.all().order_by('-date')
        else:
            return TransferRequest.objects.filter(company=user.company).order_by('-date')

class TransferRequestExcelView(LoginRequiredMixin, ListView):
    model = TransferRequest
    template_name = 'loyal_ryde_system/report_travels.html'
    context_object_name = 'transfer_requests'

    def get_queryset(self):
        queryset = TransferRequest.objects.filter(status='finalizada')
        month = self.request.GET.get('month')
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')

        if month:
            queryset = queryset.filter(date__month=month)
        if start_date and end_date:
            queryset = queryset.filter(date__range=[start_date, end_date])
        if not month and not start_date and not end_date:
            current_month = now().month
            queryset = queryset.filter(date__month=current_month)

        return queryset

    def get(self, request, *args, **kwargs):
        if request.GET.get('export') == 'true':
            return self.export_to_excel()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return self.filter_ajax()
        return super().get(request, *args, **kwargs)

    def filter_ajax(self):
        transfer_requests = self.get_queryset()
        transfer_requests_data = [
            {
                'date': tr.date.strftime('%d/%m/%Y'),
                'hour': tr.hour.strftime('%G:%i'),
                'departure_site_route': tr.departure_site_route,
                'destination_route': tr.destination_route,
                'price': tr.price,
                'deviation': {'count': tr.deviation.count()},
                'final_price': tr.final_price,
                'grafo_ceco': tr.ceco_grafo_pedido,
                'service_requested': {'company': {'name': tr.service_requested.company.name}}
            }
            for tr in transfer_requests
        ]
        return JsonResponse({'transfer_requests': transfer_requests_data})
    
    def export_to_excel(self):
        transfer_requests = self.get_queryset()

        # Obtener el mes de los parámetros de la solicitud
        month = self.request.GET.get('month')
        if month:
            month_name = format_date(datetime.strptime(month, "%m"), "MMMM", locale="es").upper()
        else:
            month_name = "MES DESCONOCIDO"

        # Crear un nuevo libro de trabajo y una hoja
        wb = Workbook()
        ws = wb.active
        ws.title = "Transfer Requests"

        # Añadir el título
        title = f"REPORTES GGMN MES DE {month_name}"
        ws.merge_cells('A1:O1')  # Unificar celdas desde A1 hasta O1
        cell = ws['A1']
        cell.value = title
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Escribir los encabezados
        headers = [
            "Fecha", "Dia", "Horas", "Pasajero", "Origen", "Destino", 
            "MONTO base $", "Horas Espera Cant.", 
            "Horas Espera $", "Desvios Cant", 
            "Desvios US$", "Lugar", 
            "Total US$", "GRAFO/CECO", "Solicitante"
        ]
        ws.append([''])  # Añadir una fila en blanco después del título
        ws.append(headers)

        # Estilo para los encabezados
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        header_border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )

        for cell in ws[3]:  # La fila de encabezados es la tercera fila
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = header_border

        # Escribir los datos
        for transfer_request in transfer_requests:
            for person in transfer_request.person_to_transfer.all():
                row = [
                    transfer_request.date.strftime('%d/%m/%Y') if transfer_request.date else '',
                    format_date(transfer_request.date, 'EEEE', locale='es') if transfer_request.date else '',
                    transfer_request.hour.strftime('%H:%M') if transfer_request.hour else '',
                    person.name if person else '',
                    transfer_request.departure_direc if transfer_request.departure_direc else '',
                    transfer_request.destination_direc if transfer_request.destination_direc else '',
                    transfer_request.price if transfer_request.price else '',
                    transfer_request.stop_time.count() if transfer_request.stop_time else '',
                    sum(stop.total_time.total_seconds() / 3600 for stop in transfer_request.stop_time.all()) * transfer_request.rate.daytime_waiting_time if transfer_request.stop_time and transfer_request.rate.daytime_waiting_time else '',
                    transfer_request.deviation.count() if transfer_request.deviation else '',
                    transfer_request.deviation.count() * transfer_request.rate.detour_local if transfer_request.deviation and transfer_request.rate.detour_local else '',
                    transfer_request.departure_direc if transfer_request.departure_direc else '',
                    transfer_request.final_price if transfer_request.final_price else '',
                    transfer_request.ceco_grafo_pedido if transfer_request.ceco_grafo_pedido else '',
                    transfer_request.service_requested.company.name if transfer_request.service_requested and transfer_request.service_requested.company else ''
                ]
                ws.append(row)

        # Ajustar el ancho de las columnas
        for col in ws.iter_cols(min_row=3, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            max_length = 0
            column = col[0].column_letter  # Obtiene la letra de la columna
            for cell in col:
                if not isinstance(cell, MergedCell):  # Omitir celdas unificadas
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Alinear el contenido de las celdas al centro
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Guardar el archivo en un objeto BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Crear la respuesta HTTP
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=solicitud_traslados_finalizados.xlsx'
        return response

class FilteredTransferRequestsView(LoginRequiredMixin, TemplateView):
    template_name = 'loyal_ryde_system/filtered_transfer_requests.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['companies'] = Company.objects.values_list('name', flat=True).distinct()

        # Intentar establecer la configuración regional a español
        try:
            locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
        except locale.Error:
            # Si falla, establecer a una configuración regional disponible
            locale.setlocale(locale.LC_TIME, 'C')

        # Generar la lista de meses en español
        context['months'] = [(i, calendar.month_name[i].capitalize()) for i in range(1, 13)]

        return context

    def post(self, request, *args, **kwargs):
        company_name = request.POST.get('company')
        status = request.POST.get('status')
        month = request.POST.get('month')
        date_range = request.POST.get('date_range')

        if company_name:
            company = Company.objects.filter(name=company_name).first()
            if company:
                transfer_requests = TransferRequest.objects.filter(company=company)

                if status and status != 'all':
                    transfer_requests = transfer_requests.filter(status=status)

                if month and month != 'all':
                    transfer_requests = transfer_requests.filter(date__month=month)

                if date_range:
                    start_date, end_date = date_range.split(' - ')
                    start_date = datetime.strptime(start_date, '%d/%m/%Y')
                    end_date = datetime.strptime(end_date, '%d/%m/%Y')
                    transfer_requests = transfer_requests.filter(date__range=(start_date, end_date))

                print(f"Transfer requests: {transfer_requests}")  # Depuración
                
                data = {
                    'name': company.name,
                    'email': company.email,
                    'rif': company.rif,
                    'address': company.address,
                    'phone': company.phone,
                    'image': company.image.url if company.image else None,
                    'transfer_requests': list(transfer_requests.values('status').annotate(count=Count('status'))),
                    'table_data': [
                        {
                            'date': tr.date.strftime('%d/%m/%Y'),
                            'hour': tr.hour.strftime('%H:%M'),  # Formato corregido
                            'departure_site_route': tr.rate.route.departure_point.name + ' - ' + tr.rate.route.departure_point.state,
                            'destination_route': tr.rate.route.arrival_point.name + ' - ' + tr.rate.route.arrival_point.state,
                            'price': tr.price,
                            'deviation_count': tr.deviation.count(),
                            'final_price': tr.final_price,
                            'payment_method': dict(payment_method_choices).get(tr.payment_method, tr.payment_method),
                            'service_requested': tr.service_requested.first_name + ' ' +tr.service_requested.last_name ,
                            'status': tr.status,
                            'id': tr.id,
                            'pdf_url': reverse('core:transfer_request_pdf', kwargs={'pk': tr.id})  # URL del PDF
                        }
                        for tr in transfer_requests
                    ]
                }
                return JsonResponse(data)
        return JsonResponse({'error': 'Company not found'}, status=404)

    def get(self, request, *args, **kwargs):
        if 'export' in request.GET:
            return self.export_to_excel(request)
        return super().get(request, *args, **kwargs)

    def export_to_excel(self, request):
        company_name = request.GET.get('company')
        status = request.GET.get('status')
        month = request.GET.get('month')
        date_range = request.GET.get('date_range')

        if company_name:
            company = Company.objects.filter(name=company_name).first()
            if company:
                transfer_requests = TransferRequest.objects.filter(company=company)

                if status and status != 'all':
                    transfer_requests = transfer_requests.filter(status=status)

                if month and month != 'all':
                    transfer_requests = transfer_requests.filter(date__month=month)

                if date_range:
                    start_date, end_date = date_range.split(' - ')
                    start_date = datetime.strptime(start_date, '%d/%m/%Y')
                    end_date = datetime.strptime(end_date, '%d/%m/%Y')
                    transfer_requests = transfer_requests.filter(date__range=(start_date, end_date))

                # Crear el archivo Excel
                wb = Workbook()
                ws = wb.active
                ws.title = "Transfer Requests"

                # Agregar encabezados
                headers = [
                    "Fecha", "Dia", "Hora", "Pasajeros", "Origen", "Destino", 
                    "MONTO base $", "Horas Espera Cant.", 
                    "Horas Espera $", "Desvios Cant", 
                    "Desvios US$", "Lugar", 
                    "Total US$", "GRAFO/CECO", "Solicitante"
                ]
                ws.append(headers)

                # Agregar datos
                for tr in transfer_requests:
                    ws.append([
                        tr.date.strftime('%d/%m/%Y'),
                        tr.date.strftime('%A'),  # Día de la semana
                        tr.hour.strftime('%H:%M'),
                        ', '.join([p.name for p in tr.person_to_transfer.all()]),  # Pasajeros
                        tr.departure_direc,
                        tr.destination_direc,
                        tr.price,
                        sum([stop.total_time.total_seconds() / 3600 for stop in tr.stop_time.all()]),  # Horas de espera
                        sum([stop.total_time.total_seconds() / 3600 * tr.rate.daytime_waiting_time for stop in tr.stop_time.all()]),  # Horas de espera $
                        tr.deviation.count(),
                        sum([d.desviation_number for d in tr.deviation.all()]),  # Desvios US$
                        tr.destination_direc,
                        tr.final_price,
                        tr.ceco_grafo_pedido,
                        tr.service_requested.company.name if tr.service_requested else ''
                    ])

                # Guardar el archivo en un objeto BytesIO
                response = BytesIO()
                wb.save(response)
                response.seek(0)

                # Crear la respuesta HTTP
                response = HttpResponse(response, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=solicitud_traslados_filtrados.xlsx'
                return response

        return JsonResponse({'error': 'Company not found'}, status=404)

class TransferRequestPDFView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        transfer_request_id = kwargs.get('pk')
        transfer_request = get_object_or_404(TransferRequest, pk=transfer_request_id)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="transfer_request_{transfer_request_id}.pdf"'

        doc = SimpleDocTemplate(response, pagesize=letter)
        elements = []

        styles = getSampleStyleSheet()
        # No redefinir estilos que ya existen
        if 'Heading1' not in styles:
            styles.add(ParagraphStyle(name='Heading1', fontSize=18, leading=22))
        if 'Heading2' not in styles:
            styles.add(ParagraphStyle(name='Heading2', fontSize=14, leading=18))

        elements.append(Paragraph("Detalles de la Solicitud de Traslado", styles['Heading1']))
        elements.append(Spacer(1, 12))

        # Información de salida y destino
        departure_destination_info = [
            ["Dirección de Salidas", transfer_request.departure_direc],
            ["Punto de Referencia de Salida", transfer_request.departure_landmark],
            ["Ruta de Destino", transfer_request.destination_route],
            ["Dirección de Destino", transfer_request.destination_direc],
            ["Punto de Referencia de Destino", transfer_request.destination_landmark]
        ]

        # Información de la tarifa
        rate_info = [
            ["Tipo de Vehículo", transfer_request.rate.type_vehicle.type],
            ["Precio de Ida", transfer_request.rate.price],
            ["Precio Ida y Vuelta", transfer_request.rate.price_round_trip],
            ["Hora de Espera Diurna", transfer_request.rate.daytime_waiting_time],
            ["Hora de Espera Nocturna", transfer_request.rate.nightly_waiting_time],
            ["Desvío Local", transfer_request.rate.detour_local]
        ]

        # Información del solicitante del servicio
        service_requested_info = [
            ["Solicitado por", f"{transfer_request.service_requested.first_name} {transfer_request.service_requested.last_name}"],
            ["Empresa", transfer_request.company.name]
        ]

        # Información del conductor
        try:
            driver_info = [
                ["Nombre del Conductor", f"{transfer_request.user_driver.user.first_name} {transfer_request.user_driver.user.last_name}"],
                ["Marca del Vehículo", transfer_request.user_driver.marca],
                ["Modelo del Vehículo", transfer_request.user_driver.model],
                ["Color del Vehículo", transfer_request.user_driver.color],
                ["Número de Placa", transfer_request.user_driver.plaque],
                ["Número de Pasajeros", transfer_request.user_driver.passengers_numbers],
                ["Tipo de Vehículo", transfer_request.user_driver.type.type]
            ]
        except:
            driver_info = []

        # Información de desviaciones
        deviation_info = []
        for deviation in transfer_request.deviation.all():
            deviation_info.append(["Dirección de Desvío", deviation.desviation_direc])
            deviation_info.append(["Número de Desvío", deviation.desviation_number])

        # Información de la transferencia
        transfer_info = [
            ["Fecha", transfer_request.date.strftime('%d/%m/%Y')],
            ["Hora", transfer_request.hour.strftime('%H:%M')],
            ["Método de Pago", transfer_request.get_payment_method_display()],
            ["CECO/GRAFO/PEDIDO", transfer_request.ceco_grafo_pedido if transfer_request.payment_method == 1 else ''],
            ["División", transfer_request.division if transfer_request.payment_method == 1 else ''],
            ["Aeropuerto", "Sí" if transfer_request.fly_checkbox else "No"],
            ["Aerolínea", transfer_request.airline if transfer_request.fly_checkbox else ''],
            ["Vuelo", transfer_request.flight if transfer_request.fly_checkbox else ''],
            ["Ruta de Vuelo", transfer_request.route_fly if transfer_request.fly_checkbox else ''],
            ["Personas a Transferir", ", ".join([f"{person.name} ({person.phone})" for person in transfer_request.person_to_transfer.all()])],
            ["Empresa", transfer_request.company],
            ["Observaciones", transfer_request.observations],
            ["Ida y Vuelta", "Sí" if transfer_request.is_round_trip else "No"],
            ["Precio", transfer_request.price],
            ["Precio Final", transfer_request.final_price]
        ]

        # Crear tablas para cada sección
        sections = [
            ("Información de Salida y Destino", departure_destination_info),
            ("Información de la Tarifa", rate_info),
            ("Información del Solicitante del Servicio", service_requested_info),
            ("Información del Conductor", driver_info),
            ("Información de Desviaciones", deviation_info),
            ("Información de la Transferencia", transfer_info)
        ]

        for title, data in sections:
            if data:  # Verificar que la sección no esté vacía
                elements.append(Paragraph(title, styles['Heading2']))
                table = Table(data, colWidths=[150, 400])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                elements.append(table)
                elements.append(Spacer(1, 12))

        # Comprobante
        if transfer_request.comprobante:
            elements.append(Paragraph("Comprobante", styles['Heading2']))
            elements.append(Image(transfer_request.comprobante.path, width=400, height=200))
            elements.append(Spacer(1, 12))

        doc.build(elements)
        return response


