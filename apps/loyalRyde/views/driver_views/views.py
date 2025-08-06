import calendar
import locale
import json
from babel.dates import format_date
from datetime import datetime
from io import BytesIO

from django.db import models
from django.db.models import Case, When, F, Q, Sum, Count
import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Count, F, Q, Sum
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, render
import base64

from django.urls import reverse_lazy
from django.utils.crypto import get_random_string
from django.views.generic import (CreateView, DeleteView, ListView, UpdateView, View)

from apps.loyalRyde.forms import *
from apps.loyalRyde.models import *

from apps.loyalRyde.views_initial import send_styled_email

#  Agregar conductor
class DriverAdd(LoginRequiredMixin, CreateView):
    model = CustomUser
    form_class = CustomUserCreationForm
    template_name = 'loyal_ryde_system/add_driver.html'
    success_url = reverse_lazy('core:driver_list')

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            random_password = get_random_string(length=10)
            user.set_password(random_password)
            user.save()

            
            # Crear instancia de CustomUserDriver y asignar valores
            fleet = FleetType.objects.get(id=request.POST.get('type')) 
            custom_user_driver = CustomUserDriver(
                user=user,
                ci=request.POST.get('ci'),
                marca=request.POST.get('marca'),
                model=request.POST.get('model'),
                color=request.POST.get('color'),
                plaque=request.POST.get('plaque'),
                passengers_numbers=request.POST.get('passengers_numbers'),
                type=fleet
            )
            custom_user_driver.save()

            send_styled_email(user, random_password)
            return HttpResponseRedirect(reverse_lazy('core:driver_list'))
        return render(request, self.template_name, {'form': form})
    
    def get_context_data(self,**kwargs):
        context = super().get_context_data(**kwargs)
        context['url_return'] = self.success_url
        type = FleetType.objects.all()
        context['type'] = type
        print("Contexto:")
        print(context)
        return context

class DriverUpdateView(LoginRequiredMixin, UpdateView):
    model = CustomUser
    form_class = CustomUserCreationForm
    template_name = 'loyal_ryde_system/add_driver.html'
    success_url = reverse_lazy('core:driver_list')

    def get_object(self, queryset=None):
        return get_object_or_404(CustomUser, pk=self.kwargs['pk'])

    def form_valid(self, form):
        user = form.save(commit=False)
        user.save()
        
        # Actualizar instancia de CustomUserDriver  
        custom_user_driver = CustomUserDriver.objects.get(user=user)
        fleet = FleetType.objects.get(id=self.request.POST.get('type'))
        custom_user_driver.ci = self.request.POST.get('ci')
        custom_user_driver.marca = self.request.POST.get('marca')
        custom_user_driver.model = self.request.POST.get('model')
        custom_user_driver.color = self.request.POST.get('color')
        custom_user_driver.plaque = self.request.POST.get('plaque')
        custom_user_driver.passengers_numbers = self.request.POST.get('passengers_numbers')
        custom_user_driver.type = fleet
        custom_user_driver.save()

        return HttpResponseRedirect(self.success_url)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        type = FleetType.objects.all()
        context['type'] = type
        context['driver'] = CustomUserDriver.objects.get(user=self.get_object())
        context['url_return'] = self.success_url
        return context

#  Listado de conductores
class DriverListView(LoginRequiredMixin, ListView):
    template_name = 'loyal_ryde_system/driver_list.html'
    model = CustomUserDriver
    context_object_name = 'drivers'

# eliminación de conductores
class DriverDeleteView(LoginRequiredMixin, DeleteView):
    model = CustomUser
    template_name = 'loyal_ryde_system/delete_departure.html'
    success_url = reverse_lazy('core:driver_list')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Conductor Eliminado Exitosamente')
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = "Eliminar Conductor"
        context['text'] = "¿Está seguro de eliminar este conductor?"
        context['url_cancel'] = reverse_lazy('core:driver_list')
        return context

#  Listado de conductores Activos 
class DriverActiveListView(LoginRequiredMixin, ListView):
    template_name = 'loyal_ryde_system/driver_list_active.html'
    model = CustomUserDriver
    context_object_name = 'drivers'

    def get_queryset(self):
        return CustomUserDriver.objects.filter(user__status='active')

#  Listado de conductores Pendientes
class DriverPendingListView(LoginRequiredMixin, ListView):
    template_name = 'loyal_ryde_system/driver_list_pending.html'
    model = CustomUserDriver
    context_object_name = 'drivers'

    def get_queryset(self):
        return CustomUserDriver.objects.filter(user__status='inactive')

# Listado de conductores Inactivos
class DriverPayrollView(LoginRequiredMixin, ListView):
    model = CustomUserDriver
    template_name = 'loyal_ryde_system/driver_payroll.html'
    context_object_name = 'drivers'

    def get_queryset(self):
        # Obtener todos los conductores
        drivers = CustomUserDriver.objects.all()

        # Anotar la cantidad de viajes por pagar y el monto pendiente a pagar
        drivers = drivers.annotate(
            trips_to_pay=Count(
                'transferrequest',
                filter=Q(transferrequest__status__in=['finalizada', 'Finalizada'], transferrequest__paid_driver=False),
                distinct=True
            ),
            pending_amount=Sum(
                Case(
                    When(
                        transferrequest__is_round_trip=True,
                        then=F('transferrequest__rate__driver_price_round_trip')
                    ),
                    default=F('transferrequest__rate__driver_price'),
                    output_field=models.DecimalField()
                ),
                filter=Q(transferrequest__status__in=['finalizada', 'Finalizada'], transferrequest__paid_driver=False),
                distinct=True
            ),
            trips_paid=Count(
                'transferrequest',
                filter=Q(transferrequest__status__in=['finalizada', 'Finalizada'], transferrequest__paid_driver=True),
                distinct=True,
            ),
            paid_amount=Sum(
                Case(
                    When(
                        transferrequest__is_round_trip=True,
                        then=F('transferrequest__rate__driver_price_round_trip')
                    ),
                    default=F('transferrequest__rate__driver_price'),
                    output_field=models.DecimalField()
                ),
                filter=Q(transferrequest__status__in=['finalizada', 'Finalizada'], transferrequest__paid_driver=True),
                distinct=True,
            ),
            
            # Cálculo del monto total por desvíos
            deviation_amount=Sum(
                F('transferrequest__rate__driver_gain_detour_local_quantity') *
                F('transferrequest__deviation__desviation_number'),
                filter=Q(
                    transferrequest__status__in=['finalizada', 'Finalizada'],
                    transferrequest__paid_driver=False,
                    transferrequest__deviation__desviation_number__gt=0
                ),
                output_field=models.DecimalField()
            )
        )

        return drivers
    

# Vista para exportar la nómina de un conductor a un archivo Excel
class DriverPayrollExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        driver_id = kwargs.get('pk')
        driver = get_object_or_404(CustomUserDriver, pk=driver_id)
        transfer_requests = TransferRequest.objects.filter(user_driver=driver, status__in=['finalizada', 'Finalizada'])

        # Crear un nuevo libro de trabajo y una hoja
        wb = Workbook()
        ws = wb.active
        ws.title = "Transfer Requests"

        # Dejar la primera fila en blanco
        ws.append([])

        # Escribir el nombre del conductor en mayúsculas y negrita en la segunda fila
        ws.merge_cells('A2:O2')
        cell = ws.cell(row=2, column=1)
        cell.value = driver.user.get_full_name().upper()
        cell.font = cell.font.copy(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Escribir los encabezados en la tercera fila
        headers = [
            "FECHA", "DIA", "HORA", "SALIDA", "DESTINO", "PASAJEROS", "EMPRESA", "SERVICIO TAXI",
            "Desvios (Precio)", "Desvios (Cantidad)", "Desvios (Total)", "Vehículo", "COSTO TOTAL SERVICIO"
        ]
        ws.append(headers)

        # Ajustar el ancho y alto de las celdas de los títulos
        thin_border = Border(left=Side(style='thin', color='000000'),
                             right=Side(style='thin', color='000000'),
                             top=Side(style='thin', color='000000'),
                             bottom=Side(style='thin', color='000000'))

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = cell.font.copy(bold=True)
            cell.border = thin_border  # Aplicar bordes a las celdas de los títulos
            ws.column_dimensions[get_column_letter(col_num)].width = 20  # Ajustar el ancho de las columnas
        ws.row_dimensions[3].height = 30  # Ajustar la altura de la fila de los títulos

        total_cost = 0

        # Escribir los datos
        for transfer_request in transfer_requests:
            rate = transfer_request.rate.driver_price_round_trip if transfer_request.is_round_trip else transfer_request.rate.driver_price
            passengers = ', '.join([str(p) for p in transfer_request.person_to_transfer.all()])
            vehicle_type = str(transfer_request.rate.type_vehicle) if transfer_request.rate and transfer_request.rate.type_vehicle else ''
            company_name = transfer_request.company.name

            # Cálculos de desvíos
            desvios_price = transfer_request.rate.driver_gain_detour_local_quantity or 0
            desvios_quantity = transfer_request.deviation.count()
            desvios_total = desvios_price * desvios_quantity

            # Costo total del servicio
            costo_total_servicio = rate + desvios_total
            total_cost += costo_total_servicio

            row = [
                transfer_request.date.strftime('%d/%m/%Y') if transfer_request.date else '',
                transfer_request.date.strftime('%A') if transfer_request.date else '',
                transfer_request.hour.strftime('%H:%M') if transfer_request.hour else '',
                transfer_request.departure_site_route if transfer_request.departure_site_route else '',
                transfer_request.destination_route if transfer_request.destination_route else '',
                passengers,
                company_name,
                rate,
                str(desvios_price) + "$",
                desvios_quantity,
                str(desvios_total) + "$",
                vehicle_type,
                str(costo_total_servicio) + "$" # COSTO TOTAL SERVICIO
            ]
            ws.append(row)

            # Aplicar bordes a las celdas de datos
            for col_num in range(1, len(row) + 1):
                cell = ws.cell(row=ws.max_row, column=col_num)
                cell.border = thin_border

        # Añadir la fila "Total a pagar"
        total_row_index = ws.max_row + 1
        ws.append([""] * 12)
        ws.merge_cells(start_row=total_row_index, start_column=1, end_row=total_row_index, end_column=11)
        cell = ws.cell(row=total_row_index, column=1)
        cell.value = "Total a pagar"
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = cell.font.copy(bold=True)
        cell.border = thin_border

        total_cell = ws.cell(row=total_row_index, column=13)
        total_cell.value = str(total_cost) + "$"
        total_cell.alignment = Alignment(horizontal='center', vertical='center')
        total_cell.font = total_cell.font.copy(bold=True)
        total_cell.border = thin_border

        # Ajustar el ancho de las columnas basado en el contenido
        for col in ws.columns:
            max_length = 0
            column = col[0].coordinate.split('1')[0]  # Obtiene la letra de la columna
            for cell in col:
                if not isinstance(cell, MergedCell):  # Ignorar celdas fusionadas
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Alinear el contenido de las celdas al centro
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Guardar el archivo en un objeto BytesIO
        response = BytesIO()
        wb.save(response)
        response.seek(0)

        # Crear la respuesta HTTP
        response = HttpResponse(response, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=nomina_conductor_{driver.user.get_full_name()}.xlsx'
        return response
    
class TransferStartView(LoginRequiredMixin, View):
    template_name = 'loyal_ryde_system/transfer_status_message.html'

    def get(self, request, *args, **kwargs):
        encoded_id = request.GET.get('id')
        if not encoded_id:
            return render(request, self.template_name, {'message': 'ID no proporcionado.'})
        try:
            decoded_id = int(base64.b64decode(encoded_id).decode())
        except Exception:
            return render(request, self.template_name, {'message': 'ID inválido.'})
        transfer = get_object_or_404(TransferRequest, pk=decoded_id)
        if transfer.status != 'aprobada':
            return render(request, self.template_name, {'message': 'No se puede iniciar el traslado. El estado debe ser "aprobada".'})
        transfer.status = 'en proceso'
        transfer.save()
        return render(request, self.template_name, {'message': 'Traslado iniciado.'})


class TransferFinishView(LoginRequiredMixin, View):
    template_name = 'loyal_ryde_system/transfer_status_message.html'

    def get(self, request, *args, **kwargs):
        encoded_id = request.GET.get('id')
        if not encoded_id:
            return render(request, self.template_name, {'message': 'ID no proporcionado.'})
        try:
            decoded_id = int(base64.b64decode(encoded_id).decode())
        except Exception:
            return render(request, self.template_name, {'message': 'ID inválido.'})
        transfer = get_object_or_404(TransferRequest, pk=decoded_id)
        if transfer.status != 'en proceso':
            return render(request, self.template_name, {'message': 'No se puede finalizar el traslado. El estado debe ser "en proceso".'})
        transfer.status = 'finalizada'
        transfer.save()
        return render(request, self.template_name, {'message': 'Traslado finalizado.'})
