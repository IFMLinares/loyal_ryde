import calendar
import locale
import json
from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.models import AnonymousUser, Group, User
from django.core import serializers
from django.core.mail import EmailMessage, EmailMultiAlternatives, send_mail
from django.db.models import Count, F, Q, Sum
from django.db.models.functions import TruncMonth
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.template.loader import render_to_string
from django.urls import reverse_lazy, reverse
from django.utils import timezone
from django.utils.crypto import get_random_string
from django.utils.dateparse import parse_date
from django.utils.html import strip_tags
from django.utils import timezone
from django.utils.timezone import now
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.views.generic import (CreateView, DeleteView, DetailView, 
                                ListView,TemplateView, UpdateView, View)

from .forms import *
from .models import *
# Create your views here.

#  INICIO
class Index(LoginRequiredMixin, TemplateView):
    template_name = 'loyal_ryde_system/index.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        now = timezone.now()
        start_of_month = now.replace(day=1)
        context['latest_transfer_requests'] = TransferRequest.objects.all().order_by('-date_created')[:10]
        context['in_progress_transfer_requests'] = TransferRequest.objects.filter(status='en proceso').order_by('-date_created')[:10]
        context['total_transfer_requests'] = TransferRequest.objects.count()
        context['total_records'] = TransferRequest.objects.filter(date_created__gte=start_of_month).count()
        return context

# VISTAS DE USUARIOS (PARA ADMINISTRADORES)

#  Agregar usuario
class UserAdd(LoginRequiredMixin, CreateView):
    model = CustomUser
    form_class = CustomUserCreationForm
    template_name = 'loyal_ryde_system/add_user.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()

        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            random_password = get_random_string(length=10)
            user.set_password(random_password)
            user.save()
            send_styled_email(user,random_password)
            if user.role == 'administrador':
                return HttpResponseRedirect(reverse_lazy('core:user_list_admin'))
            elif user.role == 'despachador':
                return HttpResponseRedirect(reverse_lazy('core:user_list_dispatcher'))
            elif user.role == 'supervisor':
                return HttpResponseRedirect(reverse_lazy('core:user_list_supervisor'))
            elif user.role == 'operator':
                return HttpResponseRedirect(reverse_lazy('core:user_list_operator'))
        return render(request, self.template_name, {'form': form})

#  Agregar usuario (de las eempresas)
class UserCompanyAdd(LoginRequiredMixin, CreateView):
    model = CustomUser
    form_class = CustomUserCreationForm
    template_name = 'loyal_ryde_system/add_user_company.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()

        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            # Verifica el nombre de la compañía antes de guardar el usuario
            if user.company.name != "Loyal ride":
                # Asigna permisos de usuario normal
                user.is_staff = False
                # Puedes agregar el usuario a un grupo con permisos específicos si es necesario
                # group = Group.objects.get(name='Normal Users')
                # user.groups.add(group)

            random_password = get_random_string(length=10)
            user.set_password(random_password)
            user.save()
            send_styled_email(user,random_password)
            return HttpResponseRedirect(reverse_lazy('core:user_list_supervisor'))
        return render(request, self.template_name, {'form': form})

# Agregar nuevo traslado
class TransferRequestCreateView(LoginRequiredMixin, CreateView):
    model = TransferRequest
    template_name = 'loyal_ryde_system/transfer_rerquest.html'
    form_class = TransferRequestForm
    success_url = reverse_lazy('core:transfer_request_list')

    def form_valid(self, form):
        # Guarda el formulario directamente
        
        form.instance.service_requested = self.request.user
        
        transfer_request = form.save()

        # Obtén el ID de la tarifa y el objeto Rates
        rate_id = form.cleaned_data['rate'].id
        rate = Rates.objects.get(id=rate_id)

        # Calcula el precio basado en si es ida y vuelta
        if form.cleaned_data['is_round_trip']:
            transfer_request.price = rate.price_round_trip
        else:
            transfer_request.price = rate.price

        # Aplica el cupón de descuento si existe

        # Calcula el precio final basado en los desvíos
        waypoints_numbers = form.cleaned_data.get('waypoints_numbers', 0)
        if waypoints_numbers > 0:
            transfer_request.final_price += (waypoints_numbers * rate.detour_local)

        # Guarda el objeto TransferRequest con los nuevos valores
        transfer_request.save()

        return super().form_valid(form)

    def post(self, request, *args, **kwargs):
        # Obtén la fecha directamente del POST
        fecha = request.POST.get('date')

        # Convierte la fecha al formato que Django espera
        fecha = datetime.strptime(fecha, '%m/%d/%Y').strftime('%Y-%m-%d')
        # Actualiza la fecha en los datos del POST
        request.POST = request.POST.copy()
        request.POST['date'] = fecha

        # Llama al método post original para guardar el TransferRequest
        response = super().post(request, *args, **kwargs)

        # Obtén el objeto TransferRequest recién creado
        transfer_request = self.object
        # Procesa los desvíos adicionales

        try:
            waypoints_numbers = int(request.POST.get('waypoints_numbers', 0))
            for i in range(3, 3 + waypoints_numbers):
                name = request.POST.get(f'waypoint-{i}')
                lat = request.POST.get(f'lat_{i}')
                lng = request.POST.get(f'lng_{i}')
                if lat and lng:
                    desviation = Desviation.objects.create(
                        desviation_direc=name,
                        desviation_number=i - 2,
                        waypoint_number=i,
                        lat=lat,
                        long=lng
                    )
                    transfer_request.deviation.add(desviation)
        except:
            pass
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        departure_points = DeparturePoint.objects.all()
        context['rates_list'] = Rates.objects.all()
        context['departure'] = departure_points
        return context

# agregar traslado (modo invitado)
class GuestTransferCreateView(CreateView):
    model = TransferRequest
    template_name = 'loyal_ryde_system/transfer_rerquest_guest.html'
    form_class = TransferRequestForm

    def form_valid(self, form):
        # Guarda el formulario directamente
        form.instance.service_requested = None
        form.instance.company = "N/A Modo Invitado"
        
        transfer_request = form.save()

        # Obtén el ID de la tarifa y el objeto Rates
        rate_id = form.cleaned_data['rate'].id
        rate = Rates.objects.get(id=rate_id)

        # Calcula el precio basado en si es ida y vuelta
        if form.cleaned_data['is_round_trip']:
            transfer_request.price = rate.price_round_trip
        else:
            transfer_request.price = rate.price

        # Calcula el precio final basado en los desvíos
        waypoints_numbers = form.cleaned_data.get('waypoints_numbers', 0)
        if waypoints_numbers > 0:
            transfer_request.final_price += (waypoints_numbers * rate.detour_local)

        # Guarda el objeto TransferRequest con los nuevos valores
        transfer_request.save()

        return HttpResponseRedirect(reverse('core:guest_transfer_success', kwargs={'pk': transfer_request.pk}))

    def post(self, request, *args, **kwargs):
        # Obtén la fecha directamente del POST
        fecha = request.POST.get('date')

        # Convierte la fecha al formato que Django espera
        fecha = datetime.strptime(fecha, '%m/%d/%Y').strftime('%Y-%m-%d')
        # Actualiza la fecha en los datos del POST
        request.POST = request.POST.copy()
        request.POST['date'] = fecha

        # Llama al método post original para guardar el TransferRequest
        response = super().post(request, *args, **kwargs)

        # Obtén el objeto TransferRequest recién creado
        transfer_request = self.object
        # Procesa los desvíos adicionales

        try:
            waypoints_numbers = int(request.POST.get('waypoints_numbers', 0))
            for i in range(3, 3 + waypoints_numbers):
                name = request.POST.get(f'waypoint-{i}')
                lat = request.POST.get(f'lat_{i}')
                lng = request.POST.get(f'lng_{i}')
                if lat and lng:
                    desviation = Desviation.objects.create(
                        desviation_direc=name,
                        desviation_number=i - 2,
                        waypoint_number=i,
                        lat=lat,
                        long=lng
                    )
                    transfer_request.deviation.add(desviation)
        except:
            pass
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        departure_points = DeparturePoint.objects.all()
        context['rates_list'] = Rates.objects.all()
        context['departure'] = departure_points
        return context

# Detalles del traslado
class TransferRequestDetailview(LoginRequiredMixin, DetailView):
    model = TransferRequest
    template_name = 'loyal_ryde_system/transfer_request_detail.html'
    context_object_name = 'detail'

    def get_context_data(self,**kwargs):
        context = super().get_context_data(**kwargs)
        departure_points = DeparturePoint.objects.all()
        context['departure'] = departure_points
        context['desviations'] = self.object.deviation.all()
        return context

class TransferRequestUpdateView(LoginRequiredMixin, UpdateView):
    model = TransferRequest
    template_name = 'loyal_ryde_system/transfer_rerquest_update.html'
    form_class = TransferRequestForm
    success_url = reverse_lazy('core:transfer_request_list')

    def form_valid(self, form):
        form.instance.service_requested = self.request.user
        messages.success(self.request, 'Formulario guardado exitosamente!')
        return super().form_valid(form)

    def post(self, request, *args, **kwargs):
        # Obtén la fecha directamente del POST
        fecha = request.POST.get('date')

        # Actualiza la fecha en los datos del POST
        request.POST = request.POST.copy()
        request.POST['date'] = fecha
        
        # Llama al método post original para guardar el TransferRequest
        response = super().post(request, *args, **kwargs)

        # Obtén el objeto TransferRequest recién creado
        transfer_request = self.object

        # Procesa los desvíos adicionales
        try:
            waypoints_numbers = int(request.POST.get('waypoints_numbers', 0))
        except ValueError:
            waypoints_numbers = 0
            print("Error: waypoints_numbers no es un entero válido")

        for i in range(3, 3 + waypoints_numbers):
            lat = request.POST.get(f'lat_{i}')
            lng = request.POST.get(f'lng_{i}')
            if lat and lng:
                try:
                    desviation = Desviation.objects.create(
                        desviation_number=i - 2,
                        waypoint_number=i,
                        lat=lat,
                        long=lng
                    )
                    transfer_request.deviation.add(desviation)
                except Exception as e:
                    print(f"Error al crear desvío {i - 2}: {e}")
            else:
                print(f"Latitud o longitud faltante para el desvío {i - 2}")

        print(transfer_request)
        return response
    
    def get_context_data(self,**kwargs):
        context = super().get_context_data(**kwargs)
        departure_points = DeparturePoint.objects.all()
        context['departure'] = departure_points
        context['desviations'] = self.object.deviation.all()
        return context

        return super().post(request, *args, **kwargs)


class GuestTransferSuccessView(TemplateView):
    template_name = 'loyal_ryde_system/guest_transfer_success.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        transfer_request_id = self.kwargs.get('pk')
        transfer_request = get_object_or_404(TransferRequest, pk=transfer_request_id)
        context['transfer_request'] = transfer_request
        return context
# Agregar Cupones de descuento

class DiscountCouponCreateView(CreateView):
    model = DiscountCoupon
    form_class = DiscountCouponForm
    template_name = 'loyal_ryde_system/add_discount_coupon.html'
    success_url = reverse_lazy('loyalRyde:discount_coupon_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        company = form.cleaned_data['company']
        users = CustomUser.objects.filter(company=company)
        for user in users:
            subject = 'New Discount Coupon Available'
            html_message = render_to_string('emails/discount_coupon.html', {'coupon': self.object})
            plain_message = strip_tags(html_message)
            send_mail(subject, plain_message, 'from@example.com', [user.email], html_message=html_message)
        return response

class DiscountCouponListView(ListView):
    model = DiscountCoupon
    template_name = 'loyal_ryde_system/discount_coupon_list.html'
    context_object_name = 'coupons'

class DiscountCouponUpdateView(UpdateView):
    model = DiscountCoupon
    form_class = DiscountCouponForm
    template_name = 'loyal_ryde_system/update_discount_coupon.html'
    success_url = reverse_lazy('discount_coupon_list')

class DiscountCouponDetailView(DetailView):
    model = DiscountCoupon
    template_name = 'loyal_ryde_system/discount_coupon_detail.html'
    context_object_name = 'coupon'
#  Agregar conductor
class DriverAdd(LoginRequiredMixin, CreateView):
    model = CustomUser
    form_class = CustomUserCreationForm
    template_name = 'loyal_ryde_system/add_driver.html'
    success_url = reverse_lazy('core:driver_list')

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            random_password = get_random_string(length=10)
            user.set_password(random_password)
            user.save()

            
            # Crear instancia de CustomUserDriver y asignar valores
            fleet = FleetType.objects.get(id=request.POST.get('type')) 
            custom_user_driver = CustomUserDriver(
                user=user,
                marca=request.POST.get('marca'),
                model=request.POST.get('model'),
                color=request.POST.get('color'),
                plaque=request.POST.get('plaque'),
                passengers_numbers=request.POST.get('passengers_numbers'),
                type=fleet
            )
            custom_user_driver.save()

            send_styled_email(user, random_password)
            return HttpResponseRedirect(reverse_lazy('core:driver_list'))
        return render(request, self.template_name, {'form': form})
    
    def get_context_data(self,**kwargs):
        context = super().get_context_data(**kwargs)
        type = FleetType.objects.all()
        context['type'] = type
        print("Contexto:")
        print(context)
        return context

#  Agregar Flota
class FleetAdd(LoginRequiredMixin, CreateView):
    model = Fleet
    form_class = AddFleetForm
    template_name = 'loyal_ryde_system/add_fleet.html'
    success_url = reverse_lazy('core:fleet_list')

#  Agregar Tipo de flota
class FleetTypeAdd(LoginRequiredMixin, CreateView):
    model = FleetType
    form_class = AddFleetTypeForm
    template_name = 'loyal_ryde_system/add_fleet_type.html'
    success_url = reverse_lazy('core:fleet_list_type')

#  Agregar Salida
class DeparturePointCreateView(LoginRequiredMixin, CreateView):
    model = DeparturePoint
    form_class = AddDepartureForm
    # fields = '__all__'
    template_name = 'loyal_ryde_system/add.html'
    success_url = reverse_lazy('core:rates_departure_list')

    def get_context_data(self,**kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = "Registrar Salida"
        return context
    
#  Agregar Destino
class ArrivalPointCreateView(LoginRequiredMixin, CreateView):
    model = ArrivalPoint
    form_class = AddArrivalForm
    # fields = '__all__'
    template_name = 'loyal_ryde_system/add.html'
    success_url = reverse_lazy('core:rates_arrival_list')

    def get_context_data(self,**kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = "Registrar Destino"
        return context

#  Agregar Ruta
class RouteCreateView(LoginRequiredMixin, CreateView):
    model = Route
    form_class = AddRouteForm
    template_name = 'loyal_ryde_system/add_route.html'
    success_url = reverse_lazy('core:route_list')

#  Agregar tarifas
class RatesCreateView(LoginRequiredMixin, CreateView):
    model = Rates
    form_class = AddRateForm
    template_name = 'loyal_ryde_system/add_rate.html'
    success_url = reverse_lazy('core:rates_list')

#  Listado de conductores
class DriverListView(LoginRequiredMixin, ListView):
    template_name = 'loyal_ryde_system/driver_list.html'
    model = CustomUserDriver
    context_object_name = 'drivers'

#  Listado de conductores Activos
class DriverActiveListView(LoginRequiredMixin, ListView):
    template_name = 'loyal_ryde_system/driver_list_active.html'
    model = CustomUserDriver
    context_object_name = 'drivers'

    def get_queryset(self):
        return CustomUserDriver.objects.filter(user__status='active')

#  Listado de conductores Pendientes
class DriverPendingListView(LoginRequiredMixin, ListView):
    template_name = 'loyal_ryde_system/driver_list_pending.html'
    model = CustomUserDriver
    context_object_name = 'drivers'

    def get_queryset(self):
        return CustomUserDriver.objects.filter(user__status='inactive')

#  Listado de Flotas
class FleetListView(LoginRequiredMixin, ListView):
    model = Fleet
    context_object_name = 'fleets'
    template_name = 'loyal_ryde_system/fleet_list.html'

#  Listado de Flotas
class FleeTypetListView(LoginRequiredMixin, ListView):
    model = FleetType
    context_object_name = 'fleets'
    template_name = 'loyal_ryde_system/fleet_list_type.html'

#  Listado de Rutas
class RouteListView(LoginRequiredMixin, ListView):
    model = Route
    context_object_name = 'routes'
    template_name = 'loyal_ryde_system/route_list.html'

#  Listado de Viajes en progreso
class TripsProgressListView(LoginRequiredMixin, ListView):
    model = TransferRequest
    context_object_name = 'transfer'
    template_name = 'loyal_ryde_system/trips_progress.html'

    def get_queryset(self):
        return TransferRequest.objects.filter(status='en proceso')

#  Listado de Viajes en completados
class TripsCompletedListView(LoginRequiredMixin, ListView):
    model = TransferRequest
    context_object_name = 'transfer'
    template_name = 'loyal_ryde_system/trips_complete.html'
    
    def get_queryset(self):
        return TransferRequest.objects.filter(status='finalizada')

#  Listado de Viajes en en espera
class TripsHoldListView(LoginRequiredMixin, ListView):
    model = TransferRequest
    context_object_name = 'transfer'
    template_name = 'loyal_ryde_system/trips_on_hold.html'
    
    def get_queryset(self):
        return TransferRequest.objects.filter(status__in=['esperando validación', 'validada'])


#  Listado de Viajes en en espera
class TripsAroveListView(LoginRequiredMixin, ListView):
    model = TransferRequest
    context_object_name = 'transfer'
    template_name = 'loyal_ryde_system/trips_approve.html'
    
    def get_queryset(self):
        return TransferRequest.objects.filter(status='aprobada')

#  Listado de Viajes programados
class TripsProgramedListView(LoginRequiredMixin, ListView):
    model = TransferRequest
    context_object_name = 'transfer'
    template_name = 'loyal_ryde_system/trips_programed.html'
    
    def get_queryset(self):
        return TransferRequest.objects.filter(status='validada')

#  Listado de Viajes cancelados
class TripsCancelledListView(LoginRequiredMixin, ListView):
    model = TransferRequest
    context_object_name = 'transfer'
    template_name = 'loyal_ryde_system/trips_canceled.html'
    
    def get_queryset(self):
        return TransferRequest.objects.filter(status='cancelada')

#  Listado de Tarifas
class RatesListView(LoginRequiredMixin, ListView):
    model = Rates
    context_object_name = 'rates'
    template_name = 'loyal_ryde_system/rates.html'

#  Listado de Destinos
class DepartureListView(LoginRequiredMixin, ListView):
    model = DeparturePoint
    context_object_name = 'departure'
    template_name = 'loyal_ryde_system/departure_list.html'

#  Listado de Salidas
class ArrivalListView(LoginRequiredMixin, ListView):
    model = ArrivalPoint
    context_object_name = 'arrival'
    template_name = 'loyal_ryde_system/arrival_list.html'

# Lista de administradores
class UserAdminListView(LoginRequiredMixin, ListView):
    model = CustomUser
    template_name = 'loyal_ryde_system/user_list_view.html'
    context_object_name = "users"

    def get_queryset(self):
        return CustomUser.objects.filter(role='administrador')

# Lista de despacahdores
class UserDispatchListView(LoginRequiredMixin, ListView):
    model = CustomUser
    template_name = 'loyal_ryde_system/user_list_view_dispatcher.html'
    context_object_name = "users"

    def get_queryset(self):
        return CustomUser.objects.filter(role='despachador')

# Lista de Operadores (Usuarios del cliente)
class UserOperatorListView(LoginRequiredMixin, ListView):
    model = CustomUser
    template_name = 'loyal_ryde_system/user_list_view_operators.html'
    context_object_name = 'users'

    def get_queryset(self):
        # Filtra el queryset para incluir solo usuarios con el rol 'operator'
        return CustomUser.objects.filter(role='operator')

# Lista de Supervisores (Usuarios del cliente)
class UserSupervisorListView(LoginRequiredMixin, ListView):
    model = CustomUser
    template_name = 'loyal_ryde_system/user_list_view_supervisors.html'
    context_object_name = 'users'

    def get_queryset(self):
        # Filtra el queryset para incluir solo usuarios con el rol 'operator'
        return CustomUser.objects.filter(role='supervisor')
# listado de Traslados
class TransferRequestListView(LoginRequiredMixin, ListView):
    model = TransferRequest
    template_name = 'loyal_ryde_system/transfer_request_list.html'
    context_object_name = 'transfer_requests'

# listado de Empresas
class CompaniesListView(LoginRequiredMixin, ListView):
    model = Company
    template_name = 'loyal_ryde_system/companies_list.html'
    context_object_name = 'companies'

#  Agregar Empresa
class CompnayAdd(LoginRequiredMixin, CreateView):
    model = Company
    form_class = AddCompanyForm
    success_url =  reverse_lazy('core:companies_list')
    template_name = 'loyal_ryde_system/add_company.html'

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Formulario guardado exitosamente!')
        return response

class TransferRequestExcelView(LoginRequiredMixin, ListView):
    model = TransferRequest
    template_name = 'loyal_ryde_system/report_travels.html'
    context_object_name = 'transfer_requests'

    def get_queryset(self):
        queryset = TransferRequest.objects.filter(status='finalizada')
        month = self.request.GET.get('month')
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')

        if month:
            queryset = queryset.filter(date__month=month)
        if start_date and end_date:
            queryset = queryset.filter(date__range=[start_date, end_date])
        if not month and not start_date and not end_date:
            current_month = now().month
            queryset = queryset.filter(date__month=current_month)

        return queryset

    def get(self, request, *args, **kwargs):
        if request.GET.get('export') == 'true':
            return self.export_to_excel()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return self.filter_ajax()
        return super().get(request, *args, **kwargs)

    def filter_ajax(self):
        transfer_requests = self.get_queryset()
        transfer_requests_data = [
            {
                'date': tr.date.strftime('%d/%m/%Y'),
                'hour': tr.hour.strftime('%G:%i'),
                'departure_site_route': tr.departure_site_route,
                'destination_route': tr.destination_route,
                'price': tr.price,
                'deviation': {'count': tr.deviation.count()},
                'final_price': tr.final_price,
                'grafo_ceco': tr.ceco_grafo_pedido,
                'service_requested': {'company': {'name': tr.service_requested.company.name}}
            }
            for tr in transfer_requests
        ]
        return JsonResponse({'transfer_requests': transfer_requests_data})

    def export_to_excel(self):
        transfer_requests = self.get_queryset()

        # Crear un nuevo libro de trabajo y una hoja
        wb = Workbook()
        ws = wb.active
        ws.title = "Transfer Requests"

        # Añadir el logo de la empresa
        logo_path = 'static/img/logo-01.png'
        img = Image(logo_path)
        img.width = 100  # Ajustar el ancho del logo
        img.height = 100  # Ajustar la altura del logo
        ws.add_image(img, 'A1')

        # Escribir los encabezados
        headers = [
            "Fecha", "Dia", "Hora", "Pasajero", "Origen", "Destino", 
            "MONTO base $", "Horas Espera Cant.", 
            "Horas Espera $", "Desvios Cant", 
            "Desvios US$", "Lugar", 
            "Total US$", "GRAFO/CECO", "Solicitante"
        ]
        ws.append([''])  # Añadir una fila en blanco al inicio
        ws.append([''])  # Añadir otra fila en blanco para el logo
        ws.append([''])  # Añadir otra fila en blanco para el logo
        ws.append([''])  # Añadir otra fila en blanco para el logo
        ws.append([''])  # Añadir otra fila en blanco para el logo
        ws.append([''])  # Añadir otra fila en blanco para el logo
        ws.append(headers)

        # Estilo para los encabezados
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        header_border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )

        for cell in ws[8]:  # La fila de encabezados es la quinta fila
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = header_border

        # Escribir los datos
        for transfer_request in transfer_requests:
            for person in transfer_request.person_to_transfer.all():
                row = [
                    transfer_request.date.strftime('%d/%m/%Y') if transfer_request.date else '',
                    transfer_request.date.strftime('%A') if transfer_request.date else '',
                    transfer_request.hour.strftime('%H:%M') if transfer_request.hour else '',
                    person.name if person else '',
                    transfer_request.departure_site_route if transfer_request.departure_site_route else '',
                    transfer_request.destination_route if transfer_request.destination_route else '',
                    transfer_request.price if transfer_request.price else '',
                    transfer_request.stop_time.count() if transfer_request.stop_time else '',
                    sum(stop.total_time.total_seconds() / 3600 for stop in transfer_request.stop_time.all()) * transfer_request.rate.daytime_waiting_time if transfer_request.stop_time and transfer_request.rate.daytime_waiting_time else '',
                    transfer_request.deviation.count() if transfer_request.deviation else '',
                    transfer_request.deviation.count() * transfer_request.rate.detour_local if transfer_request.deviation and transfer_request.rate.detour_local else '',
                    transfer_request.departure_direc if transfer_request.departure_direc else '',
                    transfer_request.final_price if transfer_request.final_price else '',
                    transfer_request.ceco_grafo_pedido if transfer_request.ceco_grafo_pedido else '',
                    transfer_request.service_requested.company.name if transfer_request.service_requested and transfer_request.service_requested.company else ''
                ]
                ws.append(row)

        # Añadir bordes a las celdas de datos
        data_border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )

        # Ajustar el ancho de las columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Obtiene la letra de la columna
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Alinear el contenido de las celdas al centro
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Guardar el archivo en un objeto BytesIO
        response = BytesIO()
        wb.save(response)
        response.seek(0)

        # Crear la respuesta HTTP
        response = HttpResponse(response, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=solicitud_traslados_finalizados.xlsx'
        return response

class GeneralReportsView(TemplateView):
    template_name = 'loyal_ryde_system/reportes_generales.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        company_name = self.request.GET.get('company_name', 'Todas')

        transfer_requests = TransferRequest.objects.all()
        if start_date and end_date:
            transfer_requests = transfer_requests.filter(date__range=[start_date, end_date])
        if company_name and company_name != 'Todas':
            transfer_requests = transfer_requests.filter(company=company_name)

        transfer_requests = list(transfer_requests.values('status').annotate(count=Count('status')))
        context['transfer_requests_json'] = json.dumps(transfer_requests)
        context['companies'] = Company.objects.all()
        context['selected_company'] = company_name
        return context

    def get(self, request, *args, **kwargs):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')
            company_name = request.GET.get('company_name', 'Todas')

            transfer_requests = TransferRequest.objects.all()
            if start_date and end_date:
                transfer_requests = transfer_requests.filter(date__range=[start_date, end_date])
            if company_name and company_name != 'Todas':
                transfer_requests = transfer_requests.filter(company=company_name)

            transfer_requests = list(transfer_requests.values('status').annotate(count=Count('status')))
            return JsonResponse({'transfer_requests_json': json.dumps(transfer_requests)})
        return super().get(request, *args, **kwargs)

class FilteredTransferRequestsView(LoginRequiredMixin, TemplateView):
    template_name = 'loyal_ryde_system/filtered_transfer_requests.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['companies'] = Company.objects.values_list('name', flat=True).distinct()
        
        # Establecer el locale a español
        locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
        
        # Generar la lista de meses en español
        context['months'] = [(i, calendar.month_name[i].capitalize()) for i in range(1, 13)]
        
        return context

    def post(self, request, *args, **kwargs):
        company_name = request.POST.get('company')
        status = request.POST.get('status')
        month = request.POST.get('month')
        date_range = request.POST.get('date_range')

        if company_name:
            company = Company.objects.filter(name=company_name).first()
            if company:
                transfer_requests = TransferRequest.objects.filter(company=company.name)

                if status and status != 'all':
                    transfer_requests = transfer_requests.filter(status=status)

                if month and month != 'all':
                    transfer_requests = transfer_requests.filter(date__month=month)

                if date_range:
                    start_date, end_date = date_range.split(' - ')
                    start_date = datetime.strptime(start_date, '%d/%m/%Y')
                    end_date = datetime.strptime(end_date, '%d/%m/%Y')
                    transfer_requests = transfer_requests.filter(date__range=(start_date, end_date))

                data = {
                    'name': company.name,
                    'email': company.email,
                    'rif': company.rif,
                    'address': company.address,
                    'phone': company.phone,
                    'image': company.image.url if company.image else None,
                    'transfer_requests': list(transfer_requests.values('status').annotate(count=Count('status'))),
                    'table_data': [
                        {
                            'date': tr.date.strftime('%d/%m/%Y'),
                            'hour': tr.hour.strftime('%H:%M'),  # Formato corregido
                            'departure_site_route': tr.departure_site_route,
                            'destination_route': tr.destination_route,
                            'price': tr.price,
                            'deviation_count': tr.deviation.count(),
                            'final_price': tr.final_price,
                            'grafo_ceco': tr.ceco_grafo_pedido,
                            'service_requested': tr.service_requested.company.name,
                            'status': tr.status,
                            'id': tr.id,
                            'pdf_url': reverse('core:transfer_request_pdf', kwargs={'pk': tr.id})  # URL del PDF
                        }
                        for tr in transfer_requests
                    ]
                }
                return JsonResponse(data)
        return JsonResponse({'error': 'Company not found'}, status=404)

    def get(self, request, *args, **kwargs):
        if 'export' in request.GET:
            return self.export_to_excel(request)
        return super().get(request, *args, **kwargs)

    def export_to_excel(self, request):
        company_name = request.GET.get('company')
        status = request.GET.get('status')
        month = request.GET.get('month')
        date_range = request.GET.get('date_range')

        if company_name:
            company = Company.objects.filter(name=company_name).first()
            if company:
                transfer_requests = TransferRequest.objects.filter(company=company.name)

                if status and status != 'all':
                    transfer_requests = transfer_requests.filter(status=status)

                if month and month != 'all':
                    transfer_requests = transfer_requests.filter(date__month=month)

                if date_range:
                    start_date, end_date = date_range.split(' - ')
                    start_date = datetime.strptime(start_date, '%d/%m/%Y')
                    end_date = datetime.strptime(end_date, '%d/%m/%Y')
                    transfer_requests = transfer_requests.filter(date__range=(start_date, end_date))

                # Crear el archivo Excel
                wb = Workbook()
                ws = wb.active
                ws.title = "Transfer Requests"

                # Agregar encabezados
                headers = [
                    "Fecha", "Dia", "Hora", "Pasajero", "Origen", "Destino", 
                    "MONTO base $", "Horas Espera Cant.", 
                    "Horas Espera $", "Desvios Cant", 
                    "Desvios US$", "Lugar", 
                    "Total US$", "GRAFO/CECO", "Solicitante"
                ]
                ws.append(headers)

                # Agregar datos
                for tr in transfer_requests:
                    ws.append([
                        tr.date.strftime('%d/%m/%Y'),
                        tr.date.strftime('%A'),  # Día de la semana
                        tr.hour.strftime('%H:%M'),
                        ', '.join([p.name for p in tr.person_to_transfer.all()]),  # Pasajeros
                        tr.departure_site_route,
                        tr.destination_route,
                        tr.price,
                        sum([stop.total_time.total_seconds() / 3600 for stop in tr.stop_time.all()]),  # Horas de espera
                        sum([stop.total_time.total_seconds() / 3600 * tr.rate.daytime_waiting_time for stop in tr.stop_time.all()]),  # Horas de espera $
                        tr.deviation.count(),
                        sum([d.desviation_number for d in tr.deviation.all()]),  # Desvios US$
                        tr.departure_site_route,
                        tr.final_price,
                        tr.ceco_grafo_pedido,
                        tr.service_requested.username if tr.service_requested else ''
                    ])

                # Guardar el archivo en un objeto BytesIO
                response = BytesIO()
                wb.save(response)
                response.seek(0)

                # Crear la respuesta HTTP
                response = HttpResponse(response, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=solicitud_traslados_filtrados.xlsx'
                return response

        return JsonResponse({'error': 'Company not found'}, status=404)

class TransferRequestPDFView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        transfer_request_id = kwargs.get('pk')
        transfer_request = get_object_or_404(TransferRequest, pk=transfer_request_id)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="transfer_request_{transfer_request_id}.pdf"'

        doc = SimpleDocTemplate(response, pagesize=letter)
        elements = []

        styles = getSampleStyleSheet()
        # No redefinir estilos que ya existen
        if 'Heading1' not in styles:
            styles.add(ParagraphStyle(name='Heading1', fontSize=18, leading=22))
        if 'Heading2' not in styles:
            styles.add(ParagraphStyle(name='Heading2', fontSize=14, leading=18))

        elements.append(Paragraph("Detalles de la Solicitud de Traslado", styles['Heading1']))
        elements.append(Spacer(1, 12))

        # Información de salida y destino
        departure_destination_info = [
            ["Dirección de Salida", transfer_request.departure_direc],
            ["Punto de Referencia de Salida", transfer_request.departure_landmark],
            ["Ruta de Destino", transfer_request.destination_route],
            ["Dirección de Destino", transfer_request.destination_direc],
            ["Punto de Referencia de Destino", transfer_request.destination_landmark]
        ]

        # Información de la tarifa
        rate_info = [
            ["Tipo de Vehículo", transfer_request.rate.type_vehicle.type],
            ["Precio de Ida", transfer_request.rate.price],
            ["Precio Ida y Vuelta", transfer_request.rate.price_round_trip],
            ["Hora de Espera Diurna", transfer_request.rate.daytime_waiting_time],
            ["Hora de Espera Nocturna", transfer_request.rate.nightly_waiting_time],
            ["Desvío Local", transfer_request.rate.detour_local]
        ]

        # Información del solicitante del servicio
        service_requested_info = [
            ["Solicitado por", f"{transfer_request.service_requested.first_name} {transfer_request.service_requested.last_name}"],
            ["Empresa", transfer_request.service_requested.company.name]
        ]

        # Información del conductor
        driver_info = [
            ["Nombre del Conductor", f"{transfer_request.user_driver.user.first_name} {transfer_request.user_driver.user.last_name}"],
            ["Marca del Vehículo", transfer_request.user_driver.marca],
            ["Modelo del Vehículo", transfer_request.user_driver.model],
            ["Color del Vehículo", transfer_request.user_driver.color],
            ["Número de Placa", transfer_request.user_driver.plaque],
            ["Número de Pasajeros", transfer_request.user_driver.passengers_numbers],
            ["Tipo de Vehículo", transfer_request.user_driver.type.type]
        ]

        # Información de desviaciones
        deviation_info = []
        for deviation in transfer_request.deviation.all():
            deviation_info.append(["Dirección de Desvío", deviation.desviation_direc])
            deviation_info.append(["Número de Desvío", deviation.desviation_number])

        # Información de la transferencia
        transfer_info = [
            ["Fecha", transfer_request.date.strftime('%d/%m/%Y')],
            ["Hora", transfer_request.hour.strftime('%H:%M')],
            ["Método de Pago", transfer_request.get_payment_method_display()],
            ["CECO/GRAFO/PEDIDO", transfer_request.ceco_grafo_pedido if transfer_request.payment_method == 1 else ''],
            ["División", transfer_request.division if transfer_request.payment_method == 1 else ''],
            ["Aeropuerto", "Sí" if transfer_request.fly_checkbox else "No"],
            ["Aerolínea", transfer_request.airline if transfer_request.fly_checkbox else ''],
            ["Vuelo", transfer_request.flight if transfer_request.fly_checkbox else ''],
            ["Ruta de Vuelo", transfer_request.route_fly if transfer_request.fly_checkbox else ''],
            ["Personas a Transferir", ", ".join([f"{person.name} ({person.phone})" for person in transfer_request.person_to_transfer.all()])],
            ["Empresa", transfer_request.company],
            ["Observaciones", transfer_request.observations],
            ["Ida y Vuelta", "Sí" if transfer_request.is_round_trip else "No"],
            ["Precio", transfer_request.price],
            ["Precio Final", transfer_request.final_price]
        ]

        # Crear tablas para cada sección
        sections = [
            ("Información de Salida y Destino", departure_destination_info),
            ("Información de la Tarifa", rate_info),
            ("Información del Solicitante del Servicio", service_requested_info),
            ("Información del Conductor", driver_info),
            ("Información de Desviaciones", deviation_info),
            ("Información de la Transferencia", transfer_info)
        ]

        for title, data in sections:
            if data:  # Verificar que la sección no esté vacía
                elements.append(Paragraph(title, styles['Heading2']))
                table = Table(data, colWidths=[150, 400])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                elements.append(table)
                elements.append(Spacer(1, 12))

        # Comprobante
        if transfer_request.comprobante:
            elements.append(Paragraph("Comprobante", styles['Heading2']))
            elements.append(Image(transfer_request.comprobante.path, width=400, height=200))
            elements.append(Spacer(1, 12))

        doc.build(elements)
        return response

class DriverPayrollView(LoginRequiredMixin, ListView):
    model = CustomUserDriver
    template_name = 'loyal_ryde_system/driver_payroll.html'
    context_object_name = 'drivers'

    def get_queryset(self):
        # Obtener todos los conductores
        drivers = CustomUserDriver.objects.all()

        # Anotar la cantidad de viajes por pagar y el monto pendiente a pagar
        drivers = drivers.annotate(
            trips_to_pay=Count('transferrequest', filter=Q(transferrequest__status__in=['finalizada', 'Finalizada'], transferrequest__paid_driver=False)),
            pending_amount=Sum(
                F('transferrequest__rate__driver_price_round_trip') if F('transferrequest__is_round_trip') else F('transferrequest__rate__driver_price'),
                filter=Q(transferrequest__status__in=['finalizada', 'Finalizada'], transferrequest__paid_driver=False)
            ),
            trips_paid=Count('transferrequest', filter=Q(transferrequest__status__in=['finalizada', 'Finalizada'], transferrequest__paid_driver=True)),
            paid_amount=Sum(
                F('transferrequest__rate__driver_price_round_trip') if F('transferrequest__is_round_trip') else F('transferrequest__rate__driver_price'),
                filter=Q(transferrequest__status__in=['finalizada', 'Finalizada'], transferrequest__paid_driver=True)
            )
        )

        return drivers

class DriverPayrollExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        driver_id = kwargs.get('pk')
        driver = get_object_or_404(CustomUserDriver, pk=driver_id)
        transfer_requests = TransferRequest.objects.filter(user_driver=driver, status__in=['finalizada', 'Finalizada'])

        # Crear un nuevo libro de trabajo y una hoja
        wb = Workbook()
        ws = wb.active
        ws.title = "Transfer Requests"

        # Dejar la primera fila en blanco
        ws.append([])

        # Escribir el nombre del conductor en mayúsculas y negrita en la segunda fila
        ws.merge_cells('A2:P2')
        cell = ws.cell(row=2, column=1)
        cell.value = driver.user.get_full_name().upper()
        cell.font = cell.font.copy(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Escribir los encabezados en la tercera fila
        headers = [
            "FECHA", "DIA", "HORA", "SALIDA", "DESTINO", "PASAJEROS", "EMPRESA", "SERVICIO TAXI", 
            "HORA ESP DIURNA", "HORA ESP NOCTUR.", "Desvios", "Monto Desvios", "Vehículo", "COSTO TOTAL SERVICIO"
        ]
        ws.append(headers)

        # Ajustar el ancho y alto de las celdas de los títulos
        thin_border = Border(left=Side(style='thin', color='000000'),
                             right=Side(style='thin', color='000000'),
                             top=Side(style='thin', color='000000'),
                             bottom=Side(style='thin', color='000000'))

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = cell.font.copy(bold=True)
            cell.border = thin_border  # Aplicar bordes a las celdas de los títulos
            ws.column_dimensions[get_column_letter(col_num)].width = 20  # Ajustar el ancho de las columnas
        ws.row_dimensions[3].height = 30  # Ajustar la altura de la fila de los títulos

        total_cost = 0

        # Escribir los datos
        for transfer_request in transfer_requests:
            rate = transfer_request.rate.driver_price_round_trip if transfer_request.is_round_trip else transfer_request.rate.driver_price
            passengers = ', '.join([str(p) for p in transfer_request.person_to_transfer.all()])
            vehicle_type = str(transfer_request.rate.type_vehicle) if transfer_request.rate and transfer_request.rate.type_vehicle else ''
            company_name = transfer_request.company
            desvios = transfer_request.deviation.count()
            monto_desvios = desvios * transfer_request.rate.detour_local
            costo_total_servicio = rate + monto_desvios
            total_cost += costo_total_servicio

            row = [
                transfer_request.date.strftime('%d/%m/%Y') if transfer_request.date else '',
                transfer_request.date.strftime('%A') if transfer_request.date else '',
                transfer_request.hour.strftime('%H:%M') if transfer_request.hour else '',
                transfer_request.departure_site_route if transfer_request.departure_site_route else '',
                transfer_request.destination_route if transfer_request.destination_route else '',
                passengers,
                company_name,
                rate,
                transfer_request.rate.daytime_waiting_time if transfer_request.rate else '',
                transfer_request.rate.nightly_waiting_time if transfer_request.rate else '',
                desvios,
                monto_desvios,
                vehicle_type,
                costo_total_servicio  # COSTO TOTAL SERVICIO
            ]
            ws.append(row)

            # Aplicar bordes a las celdas de datos
            for col_num in range(1, len(row) + 1):
                cell = ws.cell(row=ws.max_row, column=col_num)
                cell.border = thin_border

        # Añadir la fila "Total a pagar"
        total_row_index = ws.max_row + 1
        ws.append([""] * 14)
        ws.merge_cells(start_row=total_row_index, start_column=1, end_row=total_row_index, end_column=13)
        cell = ws.cell(row=total_row_index, column=1)
        cell.value = "Total a pagar"
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = cell.font.copy(bold=True)
        cell.border = thin_border

        total_cell = ws.cell(row=total_row_index, column=14)
        total_cell.value = total_cost
        total_cell.alignment = Alignment(horizontal='center', vertical='center')
        total_cell.font = total_cell.font.copy(bold=True)
        total_cell.border = thin_border

        # Ajustar el ancho de las columnas basado en el contenido
        for col in ws.columns:
            max_length = 0
            column = col[0].coordinate.split('1')[0]  # Obtiene la letra de la columna
            for cell in col:
                if not isinstance(cell, MergedCell):  # Ignorar celdas fusionadas
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Alinear el contenido de las celdas al centro
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Guardar el archivo en un objeto BytesIO
        response = BytesIO()
        wb.save(response)
        response.seek(0)

        # Crear la respuesta HTTP
        response = HttpResponse(response, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=nomina_conductor_{driver.user.get_full_name()}.xlsx'
        return response

class CompanyFilteredTransferRequestsView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'loyal_ryde_system/company_filtered_transfer_requests.html'

    def test_func(self):
        return self.request.user.role in ['supervisor', 'operator']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user_company = self.request.user.company
        transfer_requests = TransferRequest.objects.filter(company=user_company)
        context['transfer_requests'] = transfer_requests
        return context

# AJAX FUNCTIONS
@csrf_exempt
def get_people_transfer(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        phone = request.POST.get('phone')
        company = request.POST.get('company')

        person = PeopleTransfer.objects.create(name=name, phone=phone, company=company)
        data = serializers.serialize('json', [person])

        return JsonResponse({'people_transfer': data}, status=200)

@csrf_exempt
def approve_request(request):
    if request.method == 'POST':
        request_id = request.POST.get('request_id')
        transfer_request = TransferRequest.objects.get(id=request_id)
        transfer_request.status = 'validada'
        transfer_request.save()
        return JsonResponse({'status': 'success', 'message': 'La solicitud ha sido validada con éxito.'})

@csrf_exempt
def approve_request_admin(request):
    if request.method == 'POST':
        request_id = request.POST.get('request_id')
        transfer_request = TransferRequest.objects.get(id=request_id)
        if(not transfer_request.user_driver):
            return JsonResponse({'status': 'error', 'message': 'No ha seleccionado un conductor.'})
        else:
            transfer_request.status = 'aprobada'
            transfer_request.save()
            return JsonResponse({'status': 'success', 'message': 'La solicitud ha sido aprobada con éxito.'})

@csrf_exempt
def cancel_request(request):
    if request.method == 'POST':
        request_id = request.POST.get('request_id')
        transfer_request = TransferRequest.objects.get(id=request_id)
        transfer_request.status = 'cancelada'
        transfer_request.save()
        return JsonResponse({'status': 'warning', 'message': 'La solicitud ha sido cancelada.'})


def get_company_image(request):
    company_id = request.GET.get('company_id')
    company = Company.objects.get(id=company_id)
    image_url = company.image.url  # Asumiendo que tu modelo Company tiene un campo de imagen
    return JsonResponse({'image_url': image_url})

def get_routes_by_departure(request):
    if request.method == "GET":
        departure_name = request.GET.get("departure")
        routes = Route.objects.filter(departure_point__name=departure_name)
        data = [{"id": route.id, "arrival_point": route.arrival_point.name} for route in routes]
        return JsonResponse(data, safe=False)

def get_rates(request):
    if request.method == "GET":
        departure_id = request.GET.get("departure_id")
        arrival_id = request.GET.get("arrival_id")
        nro = int(request.GET.get("nro"))

        try:
            # Buscar la ruta con los IDs de salida y destino
            route = Route.objects.get(departure_point__name=departure_id, arrival_point__name=arrival_id)
            # Luego, busca la tarifa asociada a esa ruta
            rate = Rates.objects.filter(route=route)
            rate_data = []
            for n in rate:
                # if n.vehicle.passengers_numbers >= nro:
                rate_data.append({
                    "rate_id": n.id,
                    'rate_vehicle': f'{n.type_vehicle}',
                    'rate_route': F'{n.route.route_name}: {n.route.departure_point}-{n.route.arrival_point}',
                    'rate_price': n.price,
                    'rate_price_round_trip': n.price_round_trip,
                    'rate_driver_gain': n.driver_gain,
                    'rate_driver_price': n.driver_price,
                    'rate_driver_price_round_trip': n.driver_price_round_trip,
                    'rate_gain_loyal_ride': n.gain_loyal_ride,
                    'rate_gain_loyal_ride_round_trip': n.gain_loyal_ride_round_trip,
                    'rate_daytime_waiting_time': n.daytime_waiting_time,
                    'rate_nightly_waiting_time': n.nightly_waiting_time,
                    'rate_detour_local': n.detour_local,
                    # Agrega más campos según tus necesidades
                })
                print(rate_data)
            response_data = {
                "rates": rate_data
            }
            return JsonResponse(response_data)
        except (Route.DoesNotExist, Rates.DoesNotExist):
            return JsonResponse({"error": "No se encontró una tarifa para esta ruta."}, status=404)


@require_POST
def verify_discount_code(request):
    code = request.POST.get('code')
    try:
        coupon = DiscountCoupon.objects.get(code=code, expiration_date__gte=timezone.now())
        return JsonResponse({'valid': True, 'discount_value': str(coupon.discount_value), 'discount_type': coupon.discount_type})
    except DiscountCoupon.DoesNotExist:
        return JsonResponse({'valid': False, 'message': 'Código de descuento no válido o expirado.'})

# Email fuinctions

def send_styled_email(user, password):
    subject = 'Credenciales del sistema loyal Ride'
    html_content = render_to_string('account/email/email_send.html', {'user': user, 'login': reverse_lazy('core:index'), 'pass': password})
    text_content = strip_tags(html_content)  # Elimina las etiquetas HTML para el contenido de texto plano

    email = EmailMultiAlternatives(subject, text_content, 'loyalride.test@gmail.com', [user.email])
    email.attach_alternative(html_content, 'text/html')  # Adjunta el contenido HTML
    email.send()

def transfer_requests_per_month(request):
    # Obtener los datos de la consulta
    data = TransferRequest.objects.annotate(month=TruncMonth('date_created')).values('month').annotate(count=Count('id')).order_by('month')
    
    # Crear un diccionario con todos los meses y valores iniciales de cero
    response_data = {month: 0 for month in calendar.month_name if month}
    
    # Actualizar el diccionario con los datos de la consulta
    for item in data:
        response_data[item['month'].strftime('%B')] = item['count']
    
    return JsonResponse(response_data)


@csrf_exempt
def delete_coupon(request):
    if request.method == 'POST':
        coupon_id = request.POST.get('coupon_id')
        try:
            coupon = DiscountCoupon.objects.get(id=coupon_id)
            coupon.delete()
            return JsonResponse({'status': 'success', 'message': 'El cupón ha sido eliminado.'})
        except DiscountCoupon.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'El cupón no existe.'})
    return JsonResponse({'status': 'error', 'message': 'Método no permitido.'})

